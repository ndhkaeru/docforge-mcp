"""
Regression tests for no-edit round-trip fidelity fixes:

1. Sheets with BOTH a DrawingML drawing (image/chart) and comments/hyperlinks:
   the drawing relationship must be added to the existing sheet rels, the
   comments VML must not be misclassified as a drawing, and <drawing> must
   precede <legacyDrawing> in the worksheet XML.
2. gray125 pattern fill with explicit colors is preserved.
3. fitToPage print setting is preserved.
4. Per-cell protection (locked=False / hidden=True) is preserved.
5. Injected conditionalFormatting blocks appear before pageMargins
   (CT_Worksheet element order).
"""
import re
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Protection
from openpyxl.styles.colors import Color

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

from core import reconstruct_excel, serialize_excel  # noqa: E402


def _read_part(path: Path, part_name: str) -> str:
    with zipfile.ZipFile(path, "r") as zf:
        return zf.read(part_name).decode("utf-8")


def _assert_rels_complete(path: Path) -> None:
    """Every r:id referenced by a worksheet must exist in its rels part."""
    with zipfile.ZipFile(path, "r") as zf:
        names = set(zf.namelist())
        for part in [n for n in names if re.match(r"xl/worksheets/sheet\d+\.xml$", n)]:
            xml = zf.read(part).decode("utf-8")
            rids = set(re.findall(r'r:id="([^"]+)"', xml))
            rels = part.rsplit("/", 1)[0] + "/_rels/" + part.rsplit("/", 1)[1] + ".rels"
            have = (set(re.findall(r'Id="([^"]+)"', zf.read(rels).decode("utf-8")))
                    if rels in names else set())
            missing = rids - have
            assert not missing, f"{part} references missing rels: {sorted(missing)}"


def _tiny_png(path: Path) -> Path:
    from PIL import Image as PILImage
    PILImage.new("RGB", (4, 4), (255, 0, 0)).save(path)
    return path


def test_drawing_with_comments_and_hyperlinks_keeps_rels_intact(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "note"
    ws["A1"].comment = Comment("a comment", "tester")
    ws["B1"] = "link"
    ws["B1"].hyperlink = "https://example.com/"
    from openpyxl.drawing.image import Image
    ws.add_image(Image(str(_tiny_png(tmp_path / "img.png"))), "D2")
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    _assert_rels_complete(out)

    with zipfile.ZipFile(out) as zf:
        names = set(zf.namelist())
        drawings = [n for n in names if re.match(r"xl/drawings/drawing\d+\.xml$", n)]
        # exactly one DrawingML drawing, and it really is DrawingML (not VML)
        assert len(drawings) == 1
        assert zf.read(drawings[0]).lstrip().startswith(b"<wsDr")
        # the image survived
        assert any(n.startswith("xl/media/") for n in names)

    sheet_xml = _read_part(out, "xl/worksheets/sheet1.xml")
    draw_pos = sheet_xml.find("<drawing ")
    legacy_pos = sheet_xml.find("<legacyDrawing")
    assert draw_pos != -1
    if legacy_pos != -1:
        assert draw_pos < legacy_pos, "<drawing> must precede <legacyDrawing>"

    # comment and hyperlink survived
    wb2 = openpyxl.load_workbook(out)
    assert wb2.active["A1"].comment is not None
    assert wb2.active["B1"].hyperlink is not None

    # SECOND round-trip: the first output uses ../drawings/… relative rel
    # targets (like Excel-authored files) — the image must survive again.
    out2 = tmp_path / "out2.xlsx"
    reconstruct_excel(serialize_excel(str(out)), str(out2))
    _assert_rels_complete(out2)
    with zipfile.ZipFile(out2) as zf:
        assert any(n.startswith("xl/media/") for n in zf.namelist()), \
            "image lost on second round-trip"
    assert len(openpyxl.load_workbook(out2).active._images) == 1


def test_gray125_fill_with_colors_is_preserved(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "patterned"
    ws["A1"].fill = PatternFill(
        fill_type="gray125",
        fgColor=Color(rgb="FF4472C4"),
        bgColor=Color(rgb="FFFFFFFF"),
    )
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    wb2 = openpyxl.load_workbook(out)
    fill = wb2.active["A1"].fill
    assert fill.fill_type == "gray125"
    assert fill.fgColor.rgb == "FF4472C4"
    assert fill.bgColor.rgb == "FFFFFFFF"


def test_fit_to_page_is_preserved(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    wb2 = openpyxl.load_workbook(out)
    pspr = wb2.active.sheet_properties.pageSetUpPr
    assert pspr is not None and pspr.fitToPage
    assert wb2.active.page_setup.fitToWidth == 1


def test_cell_protection_is_preserved(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "locked"
    ws["B1"] = "unlocked"
    ws["B1"].protection = Protection(locked=False)
    ws["C1"] = "hidden formula"
    ws["C1"].protection = Protection(locked=True, hidden=True)
    ws.protection.sheet = True
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2.active
    assert ws2.protection.sheet
    assert ws2["A1"].protection.locked is True
    assert ws2["B1"].protection.locked is False
    assert ws2["C1"].protection.hidden is True


def test_mc_ignorable_only_lists_declared_prefixes(tmp_path):
    """Excel refuses files whose mc:Ignorable names undeclared prefixes.

    Real Excel-authored sheets carry mc:Ignorable="x14ac xr xr2 xr3" while the
    reconstructed root only declares the prefixes actually used (x14ac).
    """
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    wb.active["A1"] = "x"
    wb.save(src)

    # Patch source like a real Excel file: extra Ignorable prefixes + x14ac attr
    def patch(xml):
        xml = re.sub(
            r"<worksheet\b[^>]*>",
            ('<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
             'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
             'mc:Ignorable="x14ac xr xr2 xr3" '
             'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" '
             'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" '
             'xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" '
             'xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3">'),
            xml, count=1)
        return re.sub(
            r"<sheetFormatPr\b[^>]*/>",
            '<sheetFormatPr defaultRowHeight="15" x14ac:dyDescent="0.25"/>',
            xml, count=1)

    tmp = src.with_suffix(".tmp")
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                raw = patch(raw.decode("utf-8")).encode("utf-8")
            zout.writestr(item, raw)
    tmp.replace(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    sheet_xml = _read_part(out, "xl/worksheets/sheet1.xml")
    root = re.search(r"<worksheet\b([^>]*)>", sheet_xml).group(1)
    attrs = dict(re.findall(r'([\w:.-]+)="([^"]*)"', root))
    ignorable = attrs.get("mc:Ignorable", "")
    declared = {k[6:] for k in attrs if k.startswith("xmlns:")}
    undeclared = [t for t in ignorable.split() if t not in declared]
    assert not undeclared, f"mc:Ignorable lists undeclared prefixes: {undeclared}"
    # the x14ac extension attr must still be preserved
    assert "x14ac:dyDescent" in sheet_xml
    assert "xmlns:x14ac" in root


def test_diagonal_borders_are_preserved(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    from openpyxl.styles import Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"] = "X"
    ws["B2"].border = Border(
        diagonal=Side(style="thin", color=Color(rgb="FFFF0000")),
        diagonalUp=True, diagonalDown=True,
    )
    ws["C3"] = "Y"
    ws["C3"].border = Border(
        diagonal=Side(style="medium"), diagonalDown=True,
    )
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    wb2 = openpyxl.load_workbook(out)
    b2 = wb2.active["B2"].border
    assert b2.diagonal.border_style == "thin"
    assert b2.diagonal.color.rgb == "FFFF0000"
    assert b2.diagonalUp and b2.diagonalDown
    c3 = wb2.active["C3"].border
    assert c3.diagonal.border_style == "medium"
    assert c3.diagonalDown and not c3.diagonalUp


def test_conditional_formatting_injected_in_schema_order(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        "A1:A5",
        CellIsRule(operator="greaterThan", formula=["3"],
                   fill=PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE",
                                    fill_type="solid")),
    )
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    sheet_xml = _read_part(out, "xl/worksheets/sheet1.xml")
    cf_pos = sheet_xml.find("<conditionalFormatting")
    pm_pos = sheet_xml.find("<pageMargins")
    assert cf_pos != -1, "conditional formatting lost"
    assert pm_pos == -1 or cf_pos < pm_pos, \
        "conditionalFormatting must precede pageMargins (CT_Worksheet order)"

    # CT_Stylesheet order: injected dxfs must precede tableStyles/colors
    styles_xml = _read_part(out, "xl/styles.xml")
    dxfs_pos = styles_xml.find("<dxfs")
    assert dxfs_pos != -1, "dxfs section lost"
    for later in ("<tableStyles", "<colors"):
        lp = styles_xml.find(later)
        assert lp == -1 or dxfs_pos < lp, \
            f"dxfs must precede {later} (CT_Stylesheet order)"

    wb2 = openpyxl.load_workbook(out)
    cf_ranges = [str(r.sqref) for r in wb2.active.conditional_formatting]
    assert cf_ranges == ["A1:A5"]
