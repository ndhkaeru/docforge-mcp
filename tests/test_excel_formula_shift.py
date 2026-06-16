"""
Tests for formula reference shifting on structural edits and for the
quote-prefix input contract.

- Insert/delete rows rewrites formulas on the edited sheet AND on other
  sheets referencing it (sheet-qualified refs), plus defined names.
- References whose entire area is deleted become #REF!.
- String literals inside formulas and refs to OTHER sheets are untouched.
- Literal text starting with "=" survives round-trip as text (dt/qp markers).
- Input contract: leading apostrophe forces text + quotePrefix.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402
from core import reconstruct_excel, serialize_excel  # noqa: E402


def _load(path: Path) -> str:
    M.excel_load(str(path))
    return str(Path(path).resolve())


def _make_formula_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(2, 6):                      # B2:B5 = 1..4
        ws.cell(row=r, column=2, value=r - 1)
    ws["C1"] = "=SUM(B2:B5)"
    ws["C2"] = "=B3*2"
    ws["C3"] = '=IF(B2>0,"B2 is text here",B5)'   # literal must not shift

    other = wb.create_sheet("Other")
    other["A1"] = "=Data!B3"
    other["A2"] = "=SUM('Data'!$B$2:$B$5)"
    other["A3"] = "=Other2!B3"                 # different sheet — untouched
    wb.create_sheet("Other2")["B3"] = 99

    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names["TotalRange"] = DefinedName(
        "TotalRange", attr_text="Data!$B$2:$B$5")
    wb.save(path)


def test_insert_rows_shifts_formulas_everywhere(tmp_path):
    src = tmp_path / "f.xlsx"
    _make_formula_workbook(src)
    key = _load(src)
    rows = M.excel_clone_rows(key, "Data", 0)
    import json
    M.excel_insert_rows(key, "Data", [
        {"after_index": -1, "rows_json": json.loads(rows)},
        {"after_index": -1, "rows_json": json.loads(rows)},
    ])
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws, other = wb["Data"], wb["Other"]
    assert ws["C3"].value == "=SUM(B4:B7)"
    assert ws["C4"].value == "=B5*2"
    assert ws["C5"].value == '=IF(B4>0,"B2 is text here",B7)'  # literal intact
    assert other["A1"].value == "=Data!B5"
    assert other["A2"].value == "=SUM('Data'!$B$4:$B$7)"
    assert other["A3"].value == "=Other2!B3"                   # untouched
    assert wb.defined_names["TotalRange"].attr_text == "Data!$B$4:$B$7"


def test_delete_rows_shrinks_and_ref_errors(tmp_path):
    src = tmp_path / "f.xlsx"
    _make_formula_workbook(src)
    key = _load(src)
    M.excel_delete_rows(key, "Data", row_indices=[2])   # row 3 (B3)
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws, other = wb["Data"], wb["Other"]
    assert ws["C1"].value == "=SUM(B2:B4)"              # range shrank
    assert ws["C2"].value == "=#REF!*2"                 # single ref deleted
    assert other["A1"].value == "=Data!#REF!"
    assert wb.defined_names["TotalRange"].attr_text == "Data!$B$2:$B$4"


def test_insert_column_shifts_formulas(tmp_path):
    src = tmp_path / "f.xlsx"
    _make_formula_workbook(src)
    key = _load(src)
    M.excel_insert_column(key, "Data", after_col_index=-1)   # prepend col A
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    assert wb["Data"]["D1"].value == "=SUM(C2:C5)"
    assert wb["Other"]["A1"].value == "=Data!C3"


def test_rename_sheet_rewrites_cell_formulas(tmp_path):
    src = tmp_path / "f.xlsx"
    _make_formula_workbook(src)
    key = _load(src)
    M.excel_rename_sheet(key, "Data", "Số liệu 2026")
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    other = wb["Other"]
    assert other["A1"].value == "='Số liệu 2026'!B3"
    assert other["A2"].value == "=SUM('Số liệu 2026'!$B$2:$B$5)"
    assert other["A3"].value == "=Other2!B3"


def test_quote_prefix_input_contract(tmp_path):
    src = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "placeholder"
    wb.save(src)

    key = _load(src)
    M.excel_edit_cells(key, "Sheet", [{"row_index": 0, "edits": {
        "0": "'=not a formula",   # apostrophe → literal text
        "1": "=SUM(1,2)",         # real formula
        "2": "+tăng 5%",          # plus text — stays text automatically
        "3": "'0123",             # apostrophe-protected leading zeros
    }}])
    M.excel_save(key)
    M.excel_close(key)

    wb2 = openpyxl.load_workbook(src)
    ws = wb2.active
    assert ws["A1"].value == "=not a formula"
    assert ws["A1"].data_type == "s"
    assert bool(ws["A1"]._style and ws["A1"]._style.quotePrefix)
    assert ws["B1"].value == "=SUM(1,2)" and ws["B1"].data_type == "f"
    assert ws["C1"].value == "+tăng 5%" and ws["C1"].data_type == "s"
    assert ws["D1"].value == "0123" and ws["D1"].data_type == "s"


def test_text_looking_like_formula_roundtrips_as_text(tmp_path):
    src = tmp_path / "qp.xlsx"
    out = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    c = ws.cell(row=1, column=1, value="=this is plain text")
    c.data_type = "s"
    from openpyxl.styles.cell_style import StyleArray
    c._style = StyleArray()
    c._style.quotePrefix = 1
    ws["A2"] = "=SUM(1,2)"
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2.active
    assert ws2["A1"].value == "=this is plain text"
    assert ws2["A1"].data_type == "s", "text must not become a broken formula"
    assert bool(ws2["A1"]._style and ws2["A1"]._style.quotePrefix)
    assert ws2["A2"].data_type == "f"


def test_formula_text_not_shifted_by_structural_edit(tmp_path):
    """A dt='s' literal that LOOKS like a formula must not be rewritten."""
    src = tmp_path / "qp.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    c = ws.cell(row=5, column=1, value="=SUM(B2:B5)")    # literal text!
    c.data_type = "s"
    from openpyxl.styles.cell_style import StyleArray
    c._style = StyleArray()
    c._style.quotePrefix = 1
    wb.save(src)

    key = _load(src)
    M.excel_copy_row(key, "Data", 0, -1)   # insert a row above
    M.excel_save(key)
    M.excel_close(key)

    wb2 = openpyxl.load_workbook(src)
    assert wb2["Data"]["A6"].value == "=SUM(B2:B5)"      # content untouched
    assert wb2["Data"]["A6"].data_type == "s"
