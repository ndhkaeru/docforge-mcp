import re
import sys
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Font, GradientFill, PatternFill, Side
from openpyxl.styles.colors import Color


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

from core import reconstruct_excel, serialize_excel  # noqa: E402


THEME_TINT = "0.59999389629810485"
CUSTOM_THEME_COLOR = "123456"
SHEET_VIEW_ATTRS = (
    'workbookViewId="0" topLeftCell="C3" view="pageLayout" '
    'showGridLines="0" zoomScale="75" showRowColHeaders="0" '
    'zoomScaleNormal="90" zoomScalePageLayoutView="65"'
)


def _read_part(path: Path, part_name: str) -> str:
    with zipfile.ZipFile(path, "r") as zf:
        return zf.read(part_name).decode("utf-8")


def _rewrite_part(path: Path, part_name: str, transform) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                raw = zin.read(item.filename)
                if item.filename == part_name:
                    raw = transform(raw.decode("utf-8")).encode("utf-8")
                zout.writestr(item, raw)
    tmp.replace(path)


def _patch_sheet_xml(xml: str) -> str:
    xml = re.sub(
        r"<worksheet\b([^>]*)>",
        (
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
            'mc:Ignorable="x14ac" '
            'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac">'
        ),
        xml,
        count=1,
    )
    xml = re.sub(
        r"<sheetView\b[^>]*>",
        f"<sheetView {SHEET_VIEW_ATTRS}>",
        xml,
        count=1,
    )
    xml = re.sub(
        r"<sheetFormatPr\b[^>]*/>",
        '<sheetFormatPr defaultRowHeight="15" x14ac:dyDescent="0.25"/>',
        xml,
        count=1,
    )
    return re.sub(
        r'(<col\b[^>]*\bmin="7"[^>]*\bmax=")7("[^>]*/>)',
        r"\g<1>8\2",
        xml,
        count=1,
    )


def _make_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws["A1"] = "theme"
    ws["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor=Color(theme=9, tint=float(THEME_TINT)),
    )
    ws["A1"].font = Font(
        name="Calibri",
        size=11,
        color=Color(theme=1),
        family=2,
        scheme="minor",
    )
    ws["A1"].border = Border(bottom=Side(style="thin", color=Color(auto=True)))
    ws["B1"] = "indexed"
    ws["B1"].fill = PatternFill(
        fill_type="solid",
        fgColor=Color(indexed=5),
    )
    ws["C1"] = "gradient"
    ws["C1"].fill = GradientFill(
        type="linear",
        degree=45,
        stop=(Color(rgb="FFFF0000"), Color(rgb="FF00FF00")),
    )
    ws["D1"] = "pattern"
    ws["D1"].fill = PatternFill(
        fill_type="darkDown",
        fgColor=Color(indexed=5),
        bgColor=Color(indexed=6),
    )
    ws["H1"] = "wide"
    ws.column_dimensions["G"].width = 20
    ws.sheet_view.view = "pageLayout"
    ws.sheet_view.topLeftCell = "C3"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.sheet_view.zoomScale = 75
    ws.sheet_view.zoomScaleNormal = 90
    ws.sheet_view.zoomScalePageLayoutView = 65
    wb.save(path)

    _rewrite_part(
        path,
        "xl/styles.xml",
        lambda xml: re.sub(
            r'(<fgColor theme="9" tint=")[^"]+("/>)',
            rf"\g<1>{THEME_TINT}\2",
            xml,
            count=1,
        ),
    )
    _rewrite_part(
        path,
        "xl/theme/theme1.xml",
        lambda xml: re.sub(
            r'(<a:accent6>\s*<a:srgbClr val=")[^"]+(")',
            rf"\g<1>{CUSTOM_THEME_COLOR}\2",
            xml,
            count=1,
        ),
    )
    _rewrite_part(
        path,
        "xl/worksheets/sheet1.xml",
        _patch_sheet_xml,
    )


def test_roundtrip_preserves_raw_theme_indexed_fills_and_sheet_view_attrs(tmp_path):
    src = tmp_path / "template.xlsx"
    out = tmp_path / "roundtrip.xlsx"
    _make_template(src)

    data = serialize_excel(str(src))
    a1 = data["sheets"][0]["rows"][0]["cells"][0]
    b1 = data["sheets"][0]["rows"][0]["cells"][1]

    assert a1["fill"]
    assert data["sheets"][0]["cw"]["H"] == 20
    assert a1["_fill_raw"]["fgColor"] == {
        "type": "theme",
        "theme": 9,
        "tint": float(THEME_TINT),
    }
    assert b1["_fill_raw"]["fgColor"] == {"type": "indexed", "indexed": 5}

    reconstruct_excel(data, str(out))

    styles_xml = _read_part(out, "xl/styles.xml")
    sheet_xml = _read_part(out, "xl/worksheets/sheet1.xml")
    theme_xml = _read_part(out, "xl/theme/theme1.xml")

    assert f'<fgColor theme="9" tint="{THEME_TINT}"/>' in styles_xml
    assert '<fgColor indexed="5"/>' in styles_xml
    fills_xml = re.search(r"<fills\b[^>]*>.*?</fills>", styles_xml, re.DOTALL).group(0)
    assert '<fgColor rgb=' not in fills_xml
    assert '<gradientFill type="linear" degree="45">' in fills_xml
    assert '<patternFill patternType="darkDown"><fgColor indexed="5"/><bgColor indexed="6"/></patternFill>' in fills_xml
    assert '<color theme="1"/>' in styles_xml
    assert '<family val="2"/>' in styles_xml
    assert '<scheme val="minor"/>' in styles_xml
    assert '<color auto="1"/>' in styles_xml
    assert f"<sheetView {SHEET_VIEW_ATTRS}>" in sheet_xml
    assert '<sheetFormatPr defaultRowHeight="15" x14ac:dyDescent="0.25"/>' in sheet_xml
    assert 'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"' in sheet_xml
    assert re.search(r'<col\b(?=[^>]*\bmin="7")(?=[^>]*\bmax="8")(?=[^>]*\bwidth="20")', sheet_xml)
    assert f'<a:srgbClr val="{CUSTOM_THEME_COLOR}"/>' in theme_xml


def test_changed_fill_does_not_reuse_original_raw_fill(tmp_path):
    src = tmp_path / "template.xlsx"
    out = tmp_path / "changed.xlsx"
    _make_template(src)

    data = serialize_excel(str(src))
    data["sheets"][0]["rows"][0]["cells"][0]["fill"] = "FFFF0000"

    reconstruct_excel(data, str(out))

    styles_xml = _read_part(out, "xl/styles.xml")
    fills_xml = re.search(r"<fills\b[^>]*>.*?</fills>", styles_xml, re.DOTALL).group(0)

    assert 'rgb="FFFF0000"' in fills_xml
    assert f'<fgColor theme="9" tint="{THEME_TINT}"/>' not in fills_xml


def test_changed_column_width_does_not_reuse_original_raw_cols(tmp_path):
    src = tmp_path / "template.xlsx"
    out = tmp_path / "changed_cols.xlsx"
    _make_template(src)

    data = serialize_excel(str(src))
    data["sheets"][0]["cw"]["H"] = 22

    reconstruct_excel(data, str(out))

    sheet_xml = _read_part(out, "xl/worksheets/sheet1.xml")

    assert not re.search(r'<col\b(?=[^>]*\bmin="7")(?=[^>]*\bmax="8")(?=[^>]*\bwidth="20")', sheet_xml)
    assert re.search(r'<col\b(?=[^>]*\bmin="8")(?=[^>]*\bmax="8")(?=[^>]*\bwidth="22")', sheet_xml)
