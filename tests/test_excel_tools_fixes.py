"""
Regression tests for the MCP tool layer (main.py):

A1. Sheet-filtered load + save merges unloaded sheets back (no data loss).
A2. Macro-enabled formats are rejected at load.
A3. A failing save never corrupts the existing file (atomic write).
B1. Structural row/column edits shift all coordinate-anchored metadata:
    merges, hyperlinks, comments, data validations, conditional formatting,
    auto filter, freeze panes, hidden columns.
B2. Renaming a sheet updates defined names referencing it.
B3. Overlapping merges are rejected; editing a slave cell is rejected.
C.  Internal hyperlinks, docProps and workbook view survive a round-trip;
    private keys are stripped from read-tool output.
"""
import json
import sys
import types
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402
from core import reconstruct_excel, serialize_excel  # noqa: E402


def _make_rich_sheet(path: Path) -> None:
    """One sheet with every kind of coordinate-anchored metadata."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(1, 11):
        for c in range(1, 7):
            ws.cell(row=r, column=c, value=f"r{r}c{c}")
    ws["B5"].hyperlink = "https://example.com/"
    ws["C6"].comment = Comment("note", "tester")
    dv = DataValidation(type="list", formula1='"A,B"', allow_blank=True)
    dv.add("D2:D10")
    ws.add_data_validation(dv)
    ws.conditional_formatting.add(
        "A2:A10",
        CellIsRule(operator="equal", formula=['"x"'],
                   fill=PatternFill("solid", start_color="FFFFC7CE", end_color="FFFFC7CE")))
    ws.merge_cells("A8:B9")
    ws.auto_filter.ref = "A1:F10"
    ws.freeze_panes = "A5"
    ws.column_dimensions["E"].hidden = True
    wb.save(path)


# ── A1 ────────────────────────────────────────────────────────────────────────

def test_filtered_load_save_preserves_other_sheets(tmp_path):
    src = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "S1"
    wb.active["A1"] = "keep1"
    wb.create_sheet("S2")["A1"] = "target"
    wb.create_sheet("S3")["A1"] = "keep3"
    wb.save(src)

    key = str(src.resolve())
    M.excel_load(str(src), sheet_name="S2")
    M.excel_edit_cells(key, "S2", [{"row_index": 0, "edits": {"0": "edited"}}])
    M.excel_save(key)
    M.excel_close(key)

    wb2 = openpyxl.load_workbook(src)
    assert wb2.sheetnames == ["S1", "S2", "S3"]
    assert wb2["S1"]["A1"].value == "keep1"
    assert wb2["S2"]["A1"].value == "edited"
    assert wb2["S3"]["A1"].value == "keep3"


# ── A2 ────────────────────────────────────────────────────────────────────────

def test_macro_formats_rejected(tmp_path):
    fake = tmp_path / "book.xlsm"
    fake.write_bytes(b"PK")
    with pytest.raises(ValueError, match="not supported"):
        M.excel_load(str(fake))

def test_convert_to_markdown_without_session(tmp_path, monkeypatch):
    src = tmp_path / "simple.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Score"])
    ws.append(["Ada", 10])
    wb.save(src)

    def fake_convert_excel_to_markdown(data, *, sheet_name=None):
        assert sheet_name is None
        assert data["sheets"][0]["name"] == "Data"
        assert data["sheets"][0]["rows"][1]["cells"][0]["v"] == "Ada"
        return "# Data\n\n| Name | Score |\n| --- | --- |\n| Ada | 10 |"

    monkeypatch.setitem(
        sys.modules,
        "excel_converter",
        types.SimpleNamespace(convert_excel_to_markdown=fake_convert_excel_to_markdown),
    )

    before_sessions = dict(M._sessions)
    result = M.convert_to_markdown(str(src))

    assert result.mimeType == "text/markdown"
    assert "| Ada | 10 |" in result.text
    assert M._sessions == before_sessions

def test_convert_to_markdown_accepts_sheet_range_and_limits(tmp_path, monkeypatch):
    src = tmp_path / "limited.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["A", "B", "C"])
    ws.append([1, 2, 3])
    ws.append([4, 5, 6])
    wb.save(src)

    seen = {}

    def fake_convert_excel_to_markdown(data, *, sheet_name=None):
        seen["sheets"] = data["sheets"]
        return "ok"

    monkeypatch.setitem(
        sys.modules,
        "excel_converter",
        types.SimpleNamespace(convert_excel_to_markdown=fake_convert_excel_to_markdown),
    )

    result = M.convert_to_markdown(str(src), sheet_name="Data", range_ref="B1:C3", max_rows=2, max_cols=1)
    assert result.text == "ok"
    rows = seen["sheets"][0]["rows"]
    assert len(rows) == 2
    assert len(rows[0]["cells"]) == 1
    assert rows[0]["cells"][0]["v"] == "B"

def test_targeted_range_find_and_summary_tools(tmp_path):
    src = tmp_path / "targeted.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Score", "Formula"])
    ws.append(["Ada", 10, "=B2*2"])
    ws.append(["Grace", 20, "=B3*2"])
    ws.merge_cells("A4:B4")
    ws["A4"] = "Merged"
    wb.save(src)

    key = str(src.resolve())
    M.excel_load(str(src))
    read = json.loads(M.excel_read_range(key, "Data", range_ref="A2:B3"))
    assert read["values"] == [["Ada", 10], ["Grace", 20]]

    matches = json.loads(M.excel_find_cells(key, "Ada"))
    assert matches["count"] == 1
    assert matches["matches"][0]["row_index"] == 1
    formulas = json.loads(M.excel_find_cells(key, "B3", match_in="formula"))
    assert formulas["matches"][0]["value"] == "=B3*2"
    M.excel_close(key)

    summary = json.loads(M.excel_get_workbook_summary(str(src)))
    assert summary["sheet_count"] == 1
    assert summary["sheets"][0]["formula_count"] == 2
    assert summary["sheets"][0]["merged_ranges"] == 1

def test_excel_table_defined_name_preview_and_markdown_range(tmp_path):
    src = tmp_path / "metadata.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Score"])
    ws.append(["Ada", 10])
    ws.append(["Grace", 20])
    table = Table(displayName="Scores", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    wb.defined_names["ScoreRange"] = DefinedName("ScoreRange", attr_text="Data!$B$2:$B$3")
    wb.save(src)

    info = json.loads(M.excel_get_info(str(src)))
    assert info["sheets"][0]["table_count"] == 1

    key = str(src.resolve())
    M.excel_load(str(src))
    tables = json.loads(M.excel_list_tables(key))
    assert tables["tables"][0]["name"] == "Scores"
    names = json.loads(M.excel_list_defined_names(key))
    assert names["defined_names"][0]["name"] == "ScoreRange"
    md = M.excel_to_markdown_range(key, "Data", range_ref="A1:B2")
    assert "| Name | Score |" in md.text
    M.excel_close(key)

    preview = json.loads(M.excel_get_sheet_preview(str(src), max_rows=2, max_cols=1))
    assert preview["sheets"][0]["rows"] == [["Name"], ["Ada"]]


# ── A3 ────────────────────────────────────────────────────────────────────────

def test_failed_save_leaves_original_intact(tmp_path):
    src = tmp_path / "orig.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "precious"
    wb.save(src)
    original_bytes = src.read_bytes()

    key = M_load(src)
    data = M._get_session(key)
    data["sheets"][0]["name"] = "bad[name]"  # invalid sheet title → save fails
    with pytest.raises(Exception):
        M.excel_save(key)
    M.excel_close(key)

    assert src.read_bytes() == original_bytes, "failed save must not touch the file"


def M_load(path: Path) -> str:
    out = M.excel_load(str(path))
    return str(Path(path).resolve())


# ── B1: row insert ────────────────────────────────────────────────────────────

def test_insert_rows_shifts_anchored_metadata(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    rows = M.excel_clone_rows(key, "Data", 0)
    M.excel_insert_rows(key, "Data", [{"after_index": -1, "rows_json": json.loads(rows)},
                                      {"after_index": -1, "rows_json": json.loads(rows)}])
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws = wb["Data"]
    assert ws["B7"].hyperlink is not None            # B5 + 2 rows
    assert ws["C8"].comment is not None              # C6 + 2 rows
    dvs = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    assert dvs == ["D4:D12"]
    cf = [str(r.sqref) for r in ws.conditional_formatting]
    assert cf == ["A4:A12"]
    assert "A10:B11" in {str(r) for r in ws.merged_cells.ranges}
    assert str(ws.auto_filter.ref) == "A3:F12"
    assert ws.freeze_panes == "A7"


# ── B1: row delete ────────────────────────────────────────────────────────────

def test_delete_rows_shifts_and_shrinks(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    # delete row 0 (above) and row 2 (inside the DV/CF ranges)
    M.excel_delete_rows(key, "Data", row_indices=[0, 2])
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws = wb["Data"]
    assert ws["B3"].hyperlink is not None            # B5 − 2 rows
    assert ws["C4"].comment is not None
    dvs = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    assert dvs == ["D1:D8"]                          # D2:D10 minus two rows
    assert "A6:B7" in {str(r) for r in ws.merged_cells.ranges}


def test_delete_merge_origin_row_dissolves_cleanly(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    M.excel_delete_rows(key, "Data", row_indices=[7])   # merge origin row (A8:B9)
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws = wb["Data"]
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert "A8:B8" in merged                         # shrunk to surviving row
    # no orphan 2-row merge left behind
    assert "A8:B9" not in merged


# ── B1: column ops ────────────────────────────────────────────────────────────

def test_insert_column_shifts_anchored_metadata(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    M.excel_insert_column(key, "Data", after_col_index=-1)   # prepend column A
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws = wb["Data"]
    assert ws["C5"].hyperlink is not None            # B5 → C5
    dvs = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    assert dvs == ["E2:E10"]
    assert "B8:C9" in {str(r) for r in ws.merged_cells.ranges}
    assert ws.column_dimensions["F"].hidden          # hidden E → F


def test_delete_column_count_and_hidden_shift(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    msg = M.excel_delete_column(key, "Data", 0)
    assert "from 10 row(s)" in msg
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws = wb["Data"]
    assert ws.column_dimensions["D"].hidden          # hidden E → D
    assert ws["A5"].hyperlink is not None            # B5 → A5


# ── B2 ────────────────────────────────────────────────────────────────────────

def test_rename_sheet_updates_defined_names(tmp_path):
    src = tmp_path / "named.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Old Data"
    ws["A1"] = 1
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names["MyRange"] = DefinedName("MyRange", attr_text="'Old Data'!$A$1")
    wb.save(src)

    key = M_load(src)
    M.excel_rename_sheet(key, "Old Data", "NewData")
    M.excel_save(key)
    M.excel_close(key)

    wb2 = openpyxl.load_workbook(src)
    assert wb2.defined_names["MyRange"].attr_text == "NewData!$A$1"


# ── B3 ────────────────────────────────────────────────────────────────────────

def test_merge_overlap_rejected(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    with pytest.raises(ValueError, match="overlaps"):
        M.excel_merge_cells(key, "Data", 7, 1, 9, 2)   # overlaps A8:B9
    M.excel_close(key)


def test_edit_slave_cell_rejected(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    with pytest.raises(ValueError, match="slave"):
        M.excel_edit_cells(key, "Data", [{"row_index": 8, "edits": {"0": "x"}}])
    M.excel_close(key)


# ── C ─────────────────────────────────────────────────────────────────────────

def test_internal_hyperlink_roundtrip(tmp_path):
    src = tmp_path / "links.xlsx"
    out = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.create_sheet("Target")
    ws["A1"] = "go"
    from openpyxl.worksheet.hyperlink import Hyperlink
    ws["A1"].hyperlink = Hyperlink(ref="A1", location="Target!A1")
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))
    wb2 = openpyxl.load_workbook(out)
    hl = wb2.active["A1"].hyperlink
    assert hl is not None and hl.location == "Target!A1"


def test_docprops_and_workbook_view_roundtrip(tmp_path):
    src = tmp_path / "props.xlsx"
    out = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = 1
    wb.create_sheet("Second")["A1"] = 2
    wb.properties.creator = "Mori Daichi"
    wb.properties.title = "Spec"
    wb.properties.lastModifiedBy = "Reviewer"
    wb.active = 1  # openpyxl persists activeTab from wb.active, not views[0]
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))
    wb2 = openpyxl.load_workbook(out)
    assert wb2.properties.creator == "Mori Daichi"
    assert wb2.properties.title == "Spec"
    assert wb2.properties.lastModifiedBy == "Reviewer"
    assert wb2.views[0].activeTab == 1


def test_read_tools_strip_private_keys(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    rows = json.loads(M.excel_get_rows(key, "Data", 0, 3))
    cell = json.loads(M.excel_get_cell(key, "Data", 0, 0))
    M.excel_close(key)

    def no_private(obj):
        if isinstance(obj, dict):
            return all(not k.startswith("_") and no_private(v) for k, v in obj.items())
        if isinstance(obj, list):
            return all(no_private(x) for x in obj)
        return True

    assert no_private(rows)
    assert no_private(cell)
